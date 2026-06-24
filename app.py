from __future__ import annotations

import importlib
import hmac
import io
import json
import os
import shutil
import sys
import tempfile
import zipfile
from datetime import date
from pathlib import Path

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components


APP_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(APP_DIR))

import bmo_docling_to_excel as extractor
import real_estate_research_agent as real_estate

importlib.reload(extractor)
importlib.reload(real_estate)


st.set_page_config(
    page_title="Raman Financial Services - Bank Statement Extractor",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown(
    """
    <style>
    .block-container {max-width: 1120px; padding-top: 2rem; padding-bottom: 4rem;}
    h1, h2, h3 {letter-spacing: 0;}
    [data-testid="stMetric"] {border: 1px solid #d9dee7; padding: 14px; border-radius: 6px;}
    .status-ok {padding: 12px 14px; border-left: 4px solid #17864b; background: #eff8f3;}
    .status-error {padding: 12px 14px; border-left: 4px solid #c73838; background: #fff2f2;}
    </style>
    """,
    unsafe_allow_html=True,
)


def require_password() -> None:
    """Protect hosted deployments when APP_PASSWORD is configured."""
    expected = os.environ.get("APP_PASSWORD", "")
    if not expected:
        try:
            expected = str(st.secrets.get("APP_PASSWORD", ""))
        except FileNotFoundError:
            expected = ""
    if not expected or st.session_state.get("authenticated"):
        return

    st.title("Raman Financial Services")
    st.subheader("Bank Statement Extractor")
    st.markdown("[ramanfinancialservices.ca](https://ramanfinancialservices.ca/)")
    password = st.text_input("Password", type="password")
    if st.button("Sign in", type="primary"):
        if hmac.compare_digest(password, expected):
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    st.stop()


require_password()


SUPPORTED_BANKS = ["Auto-detect", "BMO", "CIBC", "RBC", "Tangerine", "Vancity", "TD", "Other bank"]
PAY_PERIODS = {"Weekly": 52, "Biweekly": 26, "Semi-monthly": 24, "Monthly": 12}
FED_BRACKETS_2026 = [
    (58523, 0.14),
    (117045, 0.205),
    (181440, 0.26),
    (258482, 0.29),
    (float("inf"), 0.33),
]
BC_BRACKETS_2026 = [
    (50363, 0.0506),
    (100728, 0.0770),
    (115648, 0.1050),
    (140430, 0.1229),
    (190405, 0.1470),
    (265545, 0.1680),
    (float("inf"), 0.2050),
]
FED_BPA_2026 = 16452
BC_BPA_2026 = 13216
CANADA_EMPLOYMENT_AMOUNT_2026 = 1500
BC_TAX_REDUCTION_2026 = 596
BC_TAX_REDUCTION_THRESHOLD_2026 = 24976
BC_TAX_REDUCTION_RATE_2026 = 0.0356
CPP1_RATE = 0.0595
CPP1_BASE_RATE = 0.0495
CPP2_RATE = 0.04
CPP_BASIC_EXEMPTION = 3500
CPP1_YMPE = 71300
CPP2_YAMPE = 81200
EI_RATE = 0.0163
EI_MIE = 68900
PAYROLL_COLUMNS = [
    "employee_id",
    "pay_start",
    "pay_end",
    "pay_date",
    "hours",
    "rate",
    "salary_amount",
    "overtime_hours",
    "overtime_rate",
    "overtime_pay",
    "regular_pay",
    "stat_pay",
    "sick_pay",
    "vacation_pay",
    "vac_accrual",
    "gross",
    "taxable_pay",
    "before_tax_d",
    "cpp",
    "ei",
    "tax_fed",
    "tax_prov",
    "additional_tax",
    "other_d",
    "reimb",
    "net",
    "ytd_gross",
    "ytd_cpp",
    "ytd_ei",
    "ytd_tax_fed",
    "ytd_tax_prov",
    "ytd_other_d",
    "ytd_reimb",
    "ytd_net",
    "ytd_regular_pay",
    "ytd_stat_pay",
    "ytd_sick_pay",
    "ytd_vacation_pay",
    "ytd_overtime_pay",
    "ytd_vac_accrual",
    "ytd_vac_paid",
    "pdf_link",
    "status",
]


def safe_name(value: str) -> str:
    return "".join(char if char.isalnum() or char in "-_. " else "_" for char in value).strip()


def docling_json_for_pdf(pdf_path: Path, work_dir: Path) -> Path:
    try:
        from docling.document_converter import DocumentConverter
    except ImportError as exc:
        raise RuntimeError(
            "BMO PDF conversion requires Docling. Install the website requirements and restart the app."
        ) from exc

    json_path = work_dir / f"{pdf_path.stem}.docling.json"
    result = DocumentConverter().convert(str(pdf_path))
    document = result.document
    if hasattr(document, "export_to_dict"):
        data = document.export_to_dict()
    elif hasattr(document, "model_dump"):
        data = document.model_dump()
    else:
        raise RuntimeError("This Docling version could not export JSON.")
    json_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    return json_path


def reconciliation_status(df: pd.DataFrame) -> dict:
    """Check whether extracted transaction arithmetic reaches the closing balance."""
    opening = df.loc[df["Category"] == "Opening Balance", "Balance"].dropna()
    closing = df.loc[df["Category"] == "Closing Totals", "Balance"].dropna()
    normal = df[~df["Category"].isin(["Opening Balance", "Closing Totals"])]

    if opening.empty or closing.empty:
        return {
            "status": "review",
            "message": "Opening or closing balance was not found. Compare the Excel totals with the statement.",
        }

    total_debits = float(pd.to_numeric(normal["Debit"], errors="coerce").fillna(0).sum())
    total_credits = float(pd.to_numeric(normal["Credit"], errors="coerce").fillna(0).sum())
    credit_card_statement = df["Description"].astype(str).str.contains(
        "Previous Statement Balance", case=False, regex=False
    ).any()
    if credit_card_statement:
        calculated = round(float(opening.iloc[0]) + total_debits - total_credits, 2)
    else:
        calculated = round(float(opening.iloc[0]) - total_debits + total_credits, 2)
    expected = round(float(closing.iloc[-1]), 2)
    if abs(calculated - expected) <= 0.02:
        return {
            "status": "reconciled",
            "message": f"Reconciled to the statement closing balance (${expected:,.2f}).",
        }
    return {
        "status": "review",
        "message": (
            f"Needs review: calculated closing balance is ${calculated:,.2f}, "
            f"but the statement shows ${expected:,.2f}."
        ),
    }


def process_statement(uploaded_file, selected_bank: str) -> dict:
    work_dir = Path(tempfile.mkdtemp(prefix="bank-extractor-"))
    input_path = work_dir / safe_name(uploaded_file.name)
    input_path.write_bytes(uploaded_file.getvalue())

    bank = extractor.detect_bank_from_file(input_path) if selected_bank == "Auto-detect" else selected_bank
    if bank == "Unknown":
        bank = "Other bank"

    rows = []
    direct_error = None
    if input_path.suffix.lower() == ".pdf" and bank not in {"BMO", "TD", "Other bank"}:
        try:
            rows = extractor.extract_from_statement_file(input_path, bank)
        except RuntimeError as exc:
            direct_error = exc

    if not rows:
        effective_path = input_path
        if input_path.suffix.lower() == ".pdf":
            effective_path = docling_json_for_pdf(input_path, work_dir)
        try:
            rows = extractor.extract_from_statement_file(effective_path, bank)
        except Exception as exc:
            if direct_error is not None:
                raise RuntimeError(f"Direct extraction failed: {direct_error}. Docling fallback failed: {exc}") from exc
            raise
    if not rows:
        raise RuntimeError(f"No transactions were found for {bank}.")

    df = extractor.to_dataframe(rows)
    output_name = f"{input_path.stem}_{bank}_transactions.xlsx"
    output_path = extractor.write_excel(df, work_dir / output_name)
    summary = extractor.build_summary(df)
    normal_df = df[~df["Category"].isin(["Opening Balance", "Closing Totals"])]
    visible_transactions = extractor.to_signed_amount_view(normal_df)

    metrics = {row["Metric"]: row["Value"] for _, row in summary.iterrows()}
    reconciliation = reconciliation_status(df)
    is_rbc_chequing = bank == "RBC" and not df["Description"].astype(str).str.contains(
        "Previous Statement Balance", case=False, regex=False
    ).any()
    if is_rbc_chequing and reconciliation["status"] != "reconciled":
        raise RuntimeError(
            "RBC safety check failed. The extracted transactions do not match the "
            "statement totals and closing balance, so no Excel file was produced. "
            "This statement needs review before it can be exported."
        )
    return {
        "bank": bank,
        "source": uploaded_file.name,
        "output_name": output_path.name,
        "output_bytes": output_path.read_bytes(),
        "transactions": visible_transactions,
        "summary": summary,
        "metrics": metrics,
        "reconciliation": reconciliation,
    }


def make_zip(results: list[dict]) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for result in results:
            archive.writestr(result["output_name"], result["output_bytes"])
    return buffer.getvalue()


def merge_uploaded_excels(files, output_name: str) -> tuple[bytes, pd.DataFrame, pd.DataFrame, str]:
    work_dir = Path(tempfile.mkdtemp(prefix="annual-merge-"))
    paths = []
    for file in files:
        path = work_dir / safe_name(file.name)
        path.write_bytes(file.getvalue())
        paths.append(path)

    output_name = safe_name(output_name or "Annual_transactions.xlsx")
    if not output_name.lower().endswith(".xlsx"):
        output_name += ".xlsx"
    output_path = extractor.merge_excel_files(paths, work_dir / output_name)
    transactions = pd.read_excel(output_path, sheet_name="Annual Transactions")
    summary = pd.read_excel(output_path, sheet_name="Annual Summary")
    return output_path.read_bytes(), transactions, summary, output_path.name


def merge_extracted_results(results: list[dict], output_name: str) -> tuple[bytes, pd.DataFrame, pd.DataFrame, str]:
    work_dir = Path(tempfile.mkdtemp(prefix="annual-from-extract-"))
    paths = []
    for index, result in enumerate(results, start=1):
        path = work_dir / f"{index:02d}_{safe_name(result['output_name'])}"
        path.write_bytes(result["output_bytes"])
        paths.append(path)

    output_name = safe_name(output_name or "Annual_transactions.xlsx")
    if not output_name.lower().endswith(".xlsx"):
        output_name += ".xlsx"
    output_path = extractor.merge_excel_files(paths, work_dir / output_name)
    transactions = pd.read_excel(output_path, sheet_name="Annual Transactions")
    summary = pd.read_excel(output_path, sheet_name="Annual Summary")
    return output_path.read_bytes(), transactions, summary, output_path.name


def progressive_tax(annual_income: float, brackets: list[tuple[float, float]]) -> float:
    tax = 0.0
    previous_limit = 0.0
    for limit, rate in brackets:
        taxable = min(annual_income, limit) - previous_limit
        if taxable > 0:
            tax += taxable * rate
        if annual_income <= limit:
            break
        previous_limit = limit
    return max(0.0, tax)


def calculate_payroll(values: dict) -> dict:
    periods = PAY_PERIODS[values["frequency"]]
    regular_pay = values["hours"] * values["rate"]
    overtime_pay = values["overtime_hours"] * values["overtime_rate"]
    before_tax_deductions = values.get("before_tax_deductions", 0.0)
    additional_tax = values.get("additional_tax", 0.0)
    gross = (
        regular_pay
        + values["salary_amount"]
        + overtime_pay
        + values["stat_pay"]
        + values["sick_pay"]
        + values["vacation_pay"]
        + values["bonus"]
    )

    taxable_pay = max(0.0, gross - before_tax_deductions)
    annualized_gross = taxable_pay * periods
    cpp1_base = max(0.0, min(annualized_gross, CPP1_YMPE) - CPP_BASIC_EXEMPTION)
    cpp1_base_period = cpp1_base / periods
    cpp1 = min(
        cpp1_base_period * CPP1_RATE,
        max(0.0, (CPP1_YMPE - CPP_BASIC_EXEMPTION) * CPP1_RATE - values["ytd_cpp"]),
    )
    cpp2_base = max(0.0, min(annualized_gross, CPP2_YAMPE) - CPP1_YMPE)
    cpp2 = min(
        cpp2_base * CPP2_RATE / periods,
        max(0.0, (CPP2_YAMPE - CPP1_YMPE) * CPP2_RATE - values["ytd_cpp2"]),
    )
    cpp = 0.0 if values.get("cpp_exempt") else round(max(0.0, cpp1 + cpp2), 2)
    cpp_enhanced = (
        0.0
        if values.get("cpp_exempt")
        else round(max(0.0, cpp1_base_period * (CPP1_RATE - CPP1_BASE_RATE) + cpp2), 2)
    )
    cpp_basic = 0.0 if values.get("cpp_exempt") else round(max(0.0, cpp - cpp_enhanced), 2)

    ei_max = EI_MIE * EI_RATE
    ei = 0.0 if values.get("ei_exempt") else round(min(taxable_pay * EI_RATE, max(0.0, ei_max - values["ytd_ei"])), 2)

    tax_annual_income = max(0.0, (taxable_pay - cpp_enhanced) * periods)
    federal_annual = progressive_tax(tax_annual_income, FED_BRACKETS_2026)
    federal_credit = (
        values.get("federal_claim_amount", FED_BPA_2026)
        + min(CANADA_EMPLOYMENT_AMOUNT_2026, tax_annual_income)
        + cpp_basic * periods
        + ei * periods
    ) * 0.14
    tax_fed = round(max(0.0, federal_annual - federal_credit) / periods, 2)

    bc_annual = progressive_tax(tax_annual_income, BC_BRACKETS_2026)
    bc_credit = (
        values.get("provincial_claim_amount", BC_BPA_2026)
        + cpp_basic * periods
        + ei * periods
    ) * 0.0506
    bc_reduction = max(
        0.0,
        BC_TAX_REDUCTION_2026
        - max(0.0, tax_annual_income - BC_TAX_REDUCTION_THRESHOLD_2026) * BC_TAX_REDUCTION_RATE_2026,
    )
    tax_prov = round(max(0.0, bc_annual - bc_credit - bc_reduction) / periods, 2)
    tax_fed = round(tax_fed + additional_tax, 2)

    deductions = cpp + ei + tax_fed + tax_prov + before_tax_deductions + values["other_deductions"]
    net = round(gross - deductions + values["reimbursements"], 2)
    return {
        "regular_pay": round(regular_pay, 2),
        "overtime_pay": round(overtime_pay, 2),
        "gross": round(gross, 2),
        "taxable_pay": round(taxable_pay, 2),
        "taxable_income_for_tax": round(tax_annual_income / periods, 2),
        "before_tax_deductions": round(before_tax_deductions, 2),
        "cpp": cpp,
        "cpp_basic": cpp_basic,
        "cpp_enhanced": cpp_enhanced,
        "ei": ei,
        "tax_fed": tax_fed,
        "tax_prov": tax_prov,
        "additional_tax": round(additional_tax, 2),
        "other_deductions": round(values["other_deductions"], 2),
        "reimbursements": round(values["reimbursements"], 2),
        "total_deductions": round(deductions, 2),
        "net": net,
        "employer_cpp": cpp,
        "employer_ei": round(ei * 1.4, 2),
    }


def load_payroll_register(uploaded_file) -> pd.DataFrame:
    if uploaded_file is None:
        return pd.DataFrame(columns=PAYROLL_COLUMNS)
    try:
        workbook = pd.ExcelFile(uploaded_file)
        sheet_name = "Payroll" if "Payroll" in workbook.sheet_names else workbook.sheet_names[0]
        df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
    except Exception as exc:
        raise RuntimeError(f"Could not read payroll register: {exc}") from exc
    for column in PAYROLL_COLUMNS:
        if column not in df.columns:
            df[column] = pd.NA
    return df[PAYROLL_COLUMNS]


def payroll_ytd_before(register: pd.DataFrame, employee_id: str) -> dict:
    if register.empty or not employee_id:
        base = register
    else:
        base = register[register["employee_id"].astype(str).str.strip() == str(employee_id).strip()]
    fields = [
        "gross",
        "cpp",
        "ei",
        "tax_fed",
        "tax_prov",
        "other_d",
        "reimb",
        "net",
        "regular_pay",
        "stat_pay",
        "sick_pay",
        "vacation_pay",
        "overtime_pay",
        "vac_accrual",
    ]
    return {
        field: float(pd.to_numeric(base.get(field, pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
        for field in fields
    }


def make_payroll_register_row(values: dict, calc: dict, ytd_before: dict) -> dict:
    ytd = {
        "gross": ytd_before["gross"] + calc["gross"],
        "cpp": ytd_before["cpp"] + calc["cpp"],
        "ei": ytd_before["ei"] + calc["ei"],
        "tax_fed": ytd_before["tax_fed"] + calc["tax_fed"],
        "tax_prov": ytd_before["tax_prov"] + calc["tax_prov"],
        "other_d": ytd_before["other_d"] + calc["other_deductions"],
        "reimb": ytd_before["reimb"] + calc["reimbursements"],
        "net": ytd_before["net"] + calc["net"],
        "regular_pay": ytd_before["regular_pay"] + calc["regular_pay"],
        "stat_pay": ytd_before["stat_pay"] + values["stat_pay"],
        "sick_pay": ytd_before["sick_pay"] + values["sick_pay"],
        "vacation_pay": ytd_before["vacation_pay"] + values["vacation_pay"],
        "overtime_pay": ytd_before["overtime_pay"] + calc["overtime_pay"],
        "vac_accrual": ytd_before["vac_accrual"] + values["vac_accrual"],
    }
    row = {column: pd.NA for column in PAYROLL_COLUMNS}
    row.update(
        {
            "employee_id": values["employee_id"],
            "pay_start": values["pay_start"],
            "pay_end": values["pay_end"],
            "pay_date": values["pay_date"],
            "hours": values["hours"],
            "rate": values["rate"],
            "salary_amount": values["salary_amount"],
            "overtime_hours": values["overtime_hours"],
            "overtime_rate": values["overtime_rate"],
            "overtime_pay": calc["overtime_pay"],
            "regular_pay": calc["regular_pay"],
            "stat_pay": values["stat_pay"],
            "sick_pay": values["sick_pay"],
            "vacation_pay": values["vacation_pay"],
            "vac_accrual": values["vac_accrual"],
            "gross": calc["gross"],
            "taxable_pay": calc["taxable_income_for_tax"],
            "before_tax_d": calc["before_tax_deductions"],
            "cpp": calc["cpp"],
            "ei": calc["ei"],
            "tax_fed": calc["tax_fed"],
            "tax_prov": calc["tax_prov"],
            "additional_tax": calc["additional_tax"],
            "other_d": calc["other_deductions"],
            "reimb": calc["reimbursements"],
            "net": calc["net"],
            "ytd_gross": round(ytd["gross"], 2),
            "ytd_cpp": round(ytd["cpp"], 2),
            "ytd_ei": round(ytd["ei"], 2),
            "ytd_tax_fed": round(ytd["tax_fed"], 2),
            "ytd_tax_prov": round(ytd["tax_prov"], 2),
            "ytd_other_d": round(ytd["other_d"], 2),
            "ytd_reimb": round(ytd["reimb"], 2),
            "ytd_net": round(ytd["net"], 2),
            "ytd_regular_pay": round(ytd["regular_pay"], 2),
            "ytd_stat_pay": round(ytd["stat_pay"], 2),
            "ytd_sick_pay": round(ytd["sick_pay"], 2),
            "ytd_vacation_pay": round(ytd["vacation_pay"], 2),
            "ytd_overtime_pay": round(ytd["overtime_pay"], 2),
            "ytd_vac_accrual": round(ytd["vac_accrual"], 2),
            "ytd_vac_paid": round(ytd["vacation_pay"], 2),
            "status": "Calculated",
        }
    )
    return row


def export_payroll_register(register: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        register.to_excel(writer, sheet_name="Payroll", index=False)
        summary = (
            register.groupby("employee_id", dropna=False)[["gross", "cpp", "ei", "tax_fed", "tax_prov", "net"]]
            .sum(numeric_only=True)
            .reset_index()
        )
        summary.to_excel(writer, sheet_name="Payroll Summary", index=False)
        for sheet_name, df in {"Payroll": register, "Payroll Summary": summary}.items():
            extractor.auto_adjust_columns(writer, sheet_name, df)
    return buffer.getvalue()


def build_payslip_pdf(company: dict, employee: dict, payroll: dict, calc: dict) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise RuntimeError("Payslip PDF requires reportlab. Install requirements and restart the app.") from exc

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.55 * inch,
        leftMargin=0.55 * inch,
        topMargin=0.55 * inch,
        bottomMargin=0.55 * inch,
    )
    styles = getSampleStyleSheet()
    navy = colors.HexColor("#16324F")
    teal = colors.HexColor("#087F8C")
    pale_blue = colors.HexColor("#EEF3F8")
    pale_teal = colors.HexColor("#EAF6F7")
    line = colors.HexColor("#D5DEE8")
    muted = colors.HexColor("#64748B")

    company_block = Paragraph(
        f"<font color='#FFFFFF' size='16'><b>{company['name']}</b></font><br/>"
        f"<font color='#DDE8F1' size='8'>{company['address']}<br/>{company.get('phone', '')}</font>",
        styles["Normal"],
    )
    net_label = Paragraph(
        "<para alignment='right'><font color='#D8F2F3' size='8'><b>PAY STATEMENT</b></font></para>",
        styles["Normal"],
    )
    net_amount = Paragraph(
        f"<para alignment='right' leading='22'><font color='#FFFFFF' size='20'><b>${payroll['net']:,.2f}</b></font></para>",
        styles["Normal"],
    )
    net_caption = Paragraph(
        "<para alignment='right'><font color='#D8F2F3' size='8'><b>NET PAY</b></font></para>",
        styles["Normal"],
    )
    net_block = Table([[net_label], [net_amount], [net_caption]], colWidths=[2.47 * inch])
    net_block.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 1),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ]
        )
    )
    header = Table([[company_block, net_block]], colWidths=[4.55 * inch, 2.75 * inch])
    header.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, 0), navy),
                ("BACKGROUND", (1, 0), (1, 0), teal),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 14),
                ("RIGHTPADDING", (0, 0), (-1, -1), 14),
                ("TOPPADDING", (0, 0), (-1, -1), 12),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
            ]
        )
    )

    details = [
        ["EMPLOYEE", employee["name"], "PAY PERIOD", f"{payroll['pay_start']} to {payroll['pay_end']}"],
        ["POSITION", employee["position"] or "-", "PAY DATE", str(payroll["pay_date"])],
        ["PROVINCE", employee.get("province", "British Columbia"), "FREQUENCY", payroll["frequency"]],
    ]
    details_table = Table(details, colWidths=[0.85 * inch, 2.35 * inch, 0.85 * inch, 3.25 * inch])
    details_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), pale_blue),
                ("GRID", (0, 0), (-1, -1), 0.35, line),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("TEXTCOLOR", (0, 0), (0, -1), muted),
                ("TEXTCOLOR", (2, 0), (2, -1), muted),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )

    earning_rows = [
        ("Regular pay", calc["regular_pay"]),
        ("Salary amount", payroll["salary_amount"]),
        ("Overtime pay", calc["overtime_pay"]),
        ("Stat holiday pay", payroll["stat_pay"]),
        ("Sick pay", payroll["sick_pay"]),
        ("Vacation pay", payroll["vacation_pay"]),
        ("Bonus/other taxable", payroll["bonus"]),
        ("TOTAL GROSS", calc["gross"]),
    ]
    deduction_rows = [
        ("Before-tax deductions", payroll.get("before_tax_deductions", 0.0)),
        ("CPP", payroll["cpp"]),
        ("EI", payroll["ei"]),
        ("Federal tax", payroll["tax_fed"]),
        ("Provincial tax", payroll["tax_prov"]),
        ("Other deductions", payroll["other_deductions"]),
        ("TOTAL DEDUCTIONS", payroll["total_deductions"]),
        ("NET PAY", payroll["net"]),
    ]
    comparison = [["EARNINGS", "CURRENT", "DEDUCTIONS", "CURRENT"]]
    for earning, deduction in zip(earning_rows, deduction_rows):
        comparison.append([earning[0], f"${earning[1]:,.2f}", deduction[0], f"${deduction[1]:,.2f}"])
    comparison_table = Table(comparison, colWidths=[2.25 * inch, 1.25 * inch, 2.45 * inch, 1.35 * inch])
    comparison_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (1, 0), navy),
                ("BACKGROUND", (2, 0), (3, 0), teal),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("ALIGN", (3, 0), (3, -1), "RIGHT"),
                ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (2, 1), (2, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 1), (-1, -1), 8.5),
                ("GRID", (0, 0), (-1, -1), 0.35, line),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F7F9FB")]),
                ("BACKGROUND", (0, -1), (1, -1), pale_blue),
                ("BACKGROUND", (2, -1), (3, -1), pale_teal),
                ("TEXTCOLOR", (0, -1), (1, -1), navy),
                ("TEXTCOLOR", (2, -1), (3, -1), teal),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, -1), (-1, -1), 9.5),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )

    note = Paragraph(
        "<font color='#64748B' size='8'>This statement summarizes earnings and deductions for the pay period. "
        "Review payroll deductions against CRA PDOC before remitting or filing.</font>",
        styles["Normal"],
    )
    story = [header, Spacer(1, 10), details_table, Spacer(1, 12), comparison_table, Spacer(1, 10), note]
    doc.build(story)
    return buffer.getvalue()


def build_payslip_excel(company: dict, employee: dict, payroll: dict, calc: dict) -> bytes:
    buffer = io.BytesIO()
    rows = [
        ["Company", company["name"], "", ""],
        ["Company address", company["address"], "", ""],
        ["Employee", employee["name"], "Pay date", payroll["pay_date"]],
        ["Position", employee["position"], "Pay period", f"{payroll['pay_start']} to {payroll['pay_end']}"],
        ["Frequency", payroll["frequency"], "Province", employee.get("province", "British Columbia")],
        ["", "", "", ""],
        ["Earnings", "Amount", "Deductions", "Amount"],
        ["Regular pay", calc["regular_pay"], "Before-tax deductions", payroll.get("before_tax_deductions", 0.0)],
        ["Salary amount", payroll["salary_amount"], "CPP", payroll["cpp"]],
        ["Overtime pay", calc["overtime_pay"], "EI", payroll["ei"]],
        ["Stat pay", payroll["stat_pay"], "Federal tax", payroll["tax_fed"]],
        ["Sick pay", payroll["sick_pay"], "Provincial tax", payroll["tax_prov"]],
        ["Vacation pay", payroll["vacation_pay"], "Other deductions", payroll["other_deductions"]],
        ["Bonus", payroll["bonus"], "Total deductions", payroll["total_deductions"]],
        ["Gross pay", calc["gross"], "Reimbursements", payroll["reimbursements"]],
        ["Taxable income for tax", calc.get("taxable_income_for_tax", calc["gross"]), "Net pay", payroll["net"]],
        ["", "", "", ""],
        ["Employer CPP", calc["employer_cpp"], "Employer EI", calc["employer_ei"]],
        ["Review payroll deductions against CRA PDOC before remitting or filing.", "", "", ""],
    ]
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Payslip", index=False, header=False)
        ws = writer.book["Payslip"]
        from openpyxl.styles import Alignment, Font, PatternFill

        title_fill = PatternFill("solid", fgColor="E8EEF7")
        money_format = '$#,##0.00'
        for cell in ws[7]:
            cell.font = Font(bold=True)
            cell.fill = title_fill
        for row in range(8, 17):
            ws.cell(row=row, column=2).number_format = money_format
            ws.cell(row=row, column=4).number_format = money_format
        for row in [1, 2, 3, 4, 5, 18, 19]:
            ws.cell(row=row, column=1).font = Font(bold=True)
        for column in ["A", "C"]:
            ws.column_dimensions[column].width = 26
        for column in ["B", "D"]:
            ws.column_dimensions[column].width = 18
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top")
        ws.freeze_panes = "A7"
    return buffer.getvalue()


def build_pd7a_excel(company: dict, payroll: dict, calc: dict) -> bytes:
    income_tax = round(calc["tax_fed"] + calc["tax_prov"], 2)
    employee_cpp = round(calc["cpp"], 2)
    employer_cpp = round(calc["employer_cpp"], 2)
    employee_ei = round(calc["ei"], 2)
    employer_ei = round(calc["employer_ei"], 2)
    total_cpp = round(employee_cpp + employer_cpp, 2)
    total_ei = round(employee_ei + employer_ei, 2)
    total_remittance = round(income_tax + total_cpp + total_ei, 2)
    rows = [
        ["PD7A-style payroll remittance summary", "", ""],
        ["Use this worksheet to fill CRA My Business Account or your official PD7A voucher.", "", ""],
        ["Employer name", company["name"], ""],
        ["Payroll account number", company.get("payroll_account", ""), ""],
        ["Remittance period", f"{payroll['pay_start']} to {payroll['pay_end']}", ""],
        ["Payment date", payroll["pay_date"], ""],
        ["", "", ""],
        ["Line", "Description", "Amount"],
        ["Gross payroll", "Total gross remuneration for the period", calc["gross"]],
        ["Income tax", "Federal and provincial income tax deducted", income_tax],
        ["Employee CPP", "CPP deducted from employee", employee_cpp],
        ["Employer CPP", "Employer CPP contribution", employer_cpp],
        ["Total CPP", "Employee CPP plus employer CPP", total_cpp],
        ["Employee EI", "EI deducted from employee", employee_ei],
        ["Employer EI", "Employer EI contribution", employer_ei],
        ["Total EI", "Employee EI plus employer EI", total_ei],
        ["Total remittance", "Income tax plus total CPP plus total EI", total_remittance],
        ["", "", ""],
        ["Review against CRA PDOC and CRA payroll remittance records before paying.", "", ""],
    ]
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, sheet_name="PD7A Summary", index=False, header=False)
        ws = writer.book["PD7A Summary"]
        from openpyxl.styles import Alignment, Font, PatternFill

        header_fill = PatternFill("solid", fgColor="E8EEF7")
        total_fill = PatternFill("solid", fgColor="DDEBFF")
        ws["A1"].font = Font(bold=True, size=14)
        for cell in ws[8]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for row in range(9, 18):
            ws.cell(row=row, column=3).number_format = '$#,##0.00'
        for cell in ws[17]:
            cell.font = Font(bold=True)
            cell.fill = total_fill
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 54
        ws.column_dimensions["C"].width = 18
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top")
    return buffer.getvalue()


def uploaded_enrichment_temp(uploaded_file) -> Path | None:
    if uploaded_file is None:
        return None
    payload = uploaded_file.getvalue()
    suffix = Path(uploaded_file.name).suffix.lower() or ".json"
    if suffix == ".json":
        json.loads(payload.decode("utf-8"))
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    temp.write(payload)
    temp.close()
    return Path(temp.name)


def real_estate_public_rows(rows: list[dict]) -> pd.DataFrame:
    return pd.DataFrame([{key: value for key, value in row.items() if not key.startswith("_")} for row in rows])


def real_estate_excel_bytes(tables: dict[str, pd.DataFrame]) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df in tables.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
            worksheet = writer.sheets[sheet_name[:31]]
            for column_cells in worksheet.columns:
                width = max(len(str(cell.value or "")) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 10), 44)
    return buffer.getvalue()


def uploaded_csv_temp(uploaded_file) -> Path | None:
    if uploaded_file is None:
        return None
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".csv")
    temp.write(uploaded_file.getvalue())
    temp.close()
    return Path(temp.name)


def run_real_estate_search(
    realtor_url: str,
    realtor_csv_file,
    zealty_url: str,
    zealty_file,
    rental_file,
    signal_file,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    realtor_csv_path = uploaded_csv_temp(realtor_csv_file)
    zealty_path = uploaded_enrichment_temp(zealty_file) or realtor_csv_path
    rental_path = uploaded_enrichment_temp(rental_file)
    signal_path = uploaded_enrichment_temp(signal_file)

    assumptions = real_estate.Assumptions()
    realtor = real_estate.RealtorSavedSearchProvider(
        csv_path=realtor_csv_path,
        search_url=None if realtor_csv_path else realtor_url,
    )
    zealty = real_estate.ZealtyProvider(zealty_path, zealty_url)
    rentals = real_estate.RentalProvider(rental_path)
    signals = real_estate.SignalProvider(signal_path)

    rows = []
    for listing in realtor.fetch():
        rows.append(
            real_estate.score_property(
                listing,
                zealty.fetch(listing),
                rentals.fetch(listing),
                signals.fetch(listing),
                assumptions,
            )
        )

    df = real_estate_public_rows(rows)
    top = df.sort_values("Investment Score", ascending=False).head(10) if not df.empty else df
    return df, top


def mortgage_calculator_html() -> str:
    calculator_dir = APP_DIR / "mortgage_calculator"
    html = (calculator_dir / "index.html").read_text(encoding="utf-8")
    css = (calculator_dir / "style.css").read_text(encoding="utf-8")
    js = (calculator_dir / "script.js").read_text(encoding="utf-8")
    html = html.replace('<link rel="stylesheet" href="style.css">', f"<style>{css}</style>")
    html = html.replace('<script src="script.js"></script>', f"<script>{js}</script>")
    return html


st.title("Raman Financial Services")
st.subheader("Bank Statement Extractor")
st.markdown("[ramanfinancialservices.ca](https://ramanfinancialservices.ca/)")
st.caption("Convert bank statements into reviewable Excel transactions, then combine monthly files into an annual workbook.")

extract_tab, annual_tab, payroll_tab, mortgage_tab, real_estate_tab, guide_tab = st.tabs(
    [
        "Extract statements",
        "Build annual file",
        "Payroll template",
        "Maximum Mortgage Under GDSR",
        "Real Estate Agent",
        "Guide",
    ]
)

with extract_tab:
    left, right = st.columns([1, 2])
    with left:
        selected_bank = st.selectbox("Bank", SUPPORTED_BANKS)
        uploaded_files = st.file_uploader(
            "Upload statement files",
            type=["pdf", "json"],
            accept_multiple_files=True,
            help="Upload one statement or several monthly statements.",
        )
        process_clicked = st.button("Extract transactions", type="primary", use_container_width=True)
        st.caption("Tuned: BMO, CIBC, RBC bank accounts, RBC Visa Business and Tangerine. TD and other banks use the Docling fallback.")

    with right:
        st.subheader("Processing status")
        status_placeholder = st.empty()
        if not uploaded_files:
            status_placeholder.info("Upload PDF or Docling JSON statements to begin.")

    if process_clicked:
        if not uploaded_files:
            st.error("Upload at least one statement.")
        else:
            successes = []
            failures = []
            progress = st.progress(0, text="Starting extraction...")
            for index, uploaded_file in enumerate(uploaded_files, start=1):
                try:
                    result = process_statement(uploaded_file, selected_bank)
                    successes.append(result)
                except Exception as exc:
                    failures.append({"file": uploaded_file.name, "error": str(exc)})
                progress.progress(index / len(uploaded_files), text=f"Processed {index} of {len(uploaded_files)}")
            progress.empty()

            if successes:
                st.success(f"Created {len(successes)} Excel file(s).")
                if len(successes) > 1:
                    annual_data, annual_transactions, annual_summary, annual_filename = merge_extracted_results(
                        successes,
                        "Annual_transactions.xlsx",
                    )
                    st.download_button(
                        "Download one annual workbook",
                        data=annual_data,
                        file_name=annual_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                    )
                    with st.expander("Preview annual summary"):
                        st.dataframe(annual_summary, use_container_width=True, hide_index=True)
                    st.download_button(
                        "Download monthly Excel files as ZIP",
                        data=make_zip(successes),
                        file_name="bank_statement_excels.zip",
                        mime="application/zip",
                    )

                for result in successes:
                    with st.expander(f"{result['source']} - {result['bank']}", expanded=len(successes) == 1):
                        metrics = result["metrics"]
                        cols = st.columns(4)
                        cols[0].metric("Transactions", int(metrics.get("Number of transactions", 0)))
                        cols[1].metric("Debits", f"${metrics.get('Total Debits', 0):,.2f}")
                        cols[2].metric("Credits", f"${metrics.get('Total Credits', 0):,.2f}")
                        cols[3].metric("Closing balance", f"${metrics.get('Closing Balance', 0):,.2f}")
                        reconciliation = result["reconciliation"]
                        if reconciliation["status"] == "reconciled":
                            st.success(reconciliation["message"])
                        else:
                            st.warning(reconciliation["message"])
                        st.download_button(
                            "Download Excel",
                            data=result["output_bytes"],
                            file_name=result["output_name"],
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"download-{result['output_name']}",
                        )
                        st.dataframe(result["transactions"], use_container_width=True, hide_index=True)

            if failures:
                st.error(f"{len(failures)} file(s) need review.")
                st.dataframe(pd.DataFrame(failures), use_container_width=True, hide_index=True)

with annual_tab:
    st.subheader("Combine monthly Excel files")
    monthly_files = st.file_uploader(
        "Upload monthly transaction workbooks",
        type=["xlsx"],
        accept_multiple_files=True,
        key="annual-files",
    )
    annual_name = st.text_input("Annual output name", value="Annual_transactions.xlsx")
    if st.button("Build annual workbook", type="primary"):
        if not monthly_files:
            st.error("Upload at least one monthly Excel file.")
        else:
            try:
                data, transactions, summary, filename = merge_uploaded_excels(monthly_files, annual_name)
                st.success(f"Merged {len(monthly_files)} workbook(s).")
                st.download_button(
                    "Download annual Excel",
                    data=data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                )
                st.dataframe(summary, use_container_width=True, hide_index=True)
                with st.expander("Preview annual transactions"):
                    st.dataframe(transactions, use_container_width=True, hide_index=True)
            except Exception as exc:
                st.error(str(exc))

with payroll_tab:
    st.subheader("Payroll deductions calculator")
    st.caption("Enter one pay period in a CRA PDOC-style workflow, then download a payslip PDF and updated payroll register.")
    st.info(
        "Use this like CRA PDOC: enter the employment province, pay period, taxable earnings, TD1 claim amounts, "
        "YTD CPP/EI already deducted, and any extra deductions. Review final remittances against CRA PDOC."
    )

    register_file = st.file_uploader(
        "Upload existing payroll register Excel",
        type=["xlsx"],
        key="payroll-register",
        help="Optional. If omitted, a new register will be created.",
    )
    try:
        payroll_register = load_payroll_register(register_file)
        if register_file is not None:
            st.success(f"Loaded {len(payroll_register)} existing payroll row(s).")
    except Exception as exc:
        st.error(str(exc))
        payroll_register = pd.DataFrame(columns=PAYROLL_COLUMNS)

    with st.form("payroll-form"):
        st.markdown("**Step 1 - Company**")
        company_cols = st.columns(4)
        company_name = company_cols[0].text_input("Company name", value="Raman Tax & Accounting Inc.")
        company_address = company_cols[1].text_input("Company address", value="Surrey, BC")
        company_phone = company_cols[2].text_input("Company phone")
        payroll_account = company_cols[3].text_input("Payroll account number", placeholder="123456789RP0001")

        st.markdown("**Step 2 - Employee**")
        emp_cols = st.columns(4)
        employee_name = emp_cols[0].text_input("Employee name")
        employee_id = emp_cols[1].text_input("Employee ID")
        position = emp_cols[2].text_input("Position")
        province = emp_cols[3].selectbox("Province of employment", ["British Columbia"], index=0)
        emp_extra_cols = st.columns(2)
        employee_address = emp_extra_cols[0].text_input("Employee address")
        hire_date = emp_extra_cols[1].date_input("Date of hire", value=date.today())

        st.markdown("**Step 3 - Pay period**")
        period_cols = st.columns(4)
        frequency = period_cols[0].selectbox("Pay frequency", list(PAY_PERIODS.keys()), index=1)
        pay_start = period_cols[1].date_input("Pay start", value=date.today())
        pay_end = period_cols[2].date_input("Pay end", value=date.today())
        pay_date = period_cols[3].date_input("Pay date", value=date.today())

        st.markdown("**Step 4 - Earnings**")
        earn_cols = st.columns(4)
        hours = earn_cols[0].number_input("Regular hours", min_value=0.0, value=0.0, step=0.5)
        rate = earn_cols[1].number_input("Hourly rate", min_value=0.0, value=0.0, step=0.5)
        overtime_hours = earn_cols[2].number_input("Overtime hours", min_value=0.0, value=0.0, step=0.5)
        overtime_rate = earn_cols[3].number_input("Overtime rate", min_value=0.0, value=0.0, step=0.5)
        earn_cols2 = st.columns(4)
        stat_pay = earn_cols2[0].number_input("Stat pay", min_value=0.0, value=0.0, step=10.0)
        vacation_pay = earn_cols2[1].number_input("Vacation pay paid", min_value=0.0, value=0.0, step=10.0)
        bonus = earn_cols2[2].number_input("Bonus/other taxable pay", min_value=0.0, value=0.0, step=10.0)
        reimbursements = earn_cols2[3].number_input("Reimbursements", min_value=0.0, value=0.0, step=10.0)
        earn_cols3 = st.columns(3)
        salary_amount = earn_cols3[0].number_input("Salary amount", min_value=0.0, value=0.0, step=10.0)
        sick_pay = earn_cols3[1].number_input("Sick pay", min_value=0.0, value=0.0, step=10.0)
        vac_accrual = earn_cols3[2].number_input("Vacation accrual", min_value=0.0, value=0.0, step=10.0)

        st.markdown("**Step 5 - Tax credits and exemptions**")
        claim_cols = st.columns(4)
        federal_claim_amount = claim_cols[0].number_input(
            "Federal TD1 claim amount", min_value=0.0, value=float(FED_BPA_2026), step=100.0
        )
        provincial_claim_amount = claim_cols[1].number_input(
            "BC TD1 claim amount", min_value=0.0, value=float(BC_BPA_2026), step=100.0
        )
        cpp_exempt = claim_cols[2].checkbox("CPP exempt")
        ei_exempt = claim_cols[3].checkbox("EI exempt")

        st.markdown("**Step 6 - Year-to-date and adjustments**")
        ytd_cols = st.columns(3)
        ytd_cpp = ytd_cols[0].number_input("YTD CPP already deducted", min_value=0.0, value=0.0, step=10.0)
        ytd_cpp2 = ytd_cols[1].number_input("YTD CPP2 already deducted", min_value=0.0, value=0.0, step=10.0)
        ytd_ei = ytd_cols[2].number_input("YTD EI already deducted", min_value=0.0, value=0.0, step=10.0)
        adjust_cols = st.columns(3)
        before_tax_deductions = adjust_cols[0].number_input(
            "Before-tax deductions", min_value=0.0, value=0.0, step=10.0, help="Examples: RRSP or pension amounts deducted before tax."
        )
        additional_tax = adjust_cols[1].number_input(
            "Additional tax to deduct", min_value=0.0, value=0.0, step=10.0, help="Extra income tax requested by the employee."
        )
        other_deductions = adjust_cols[2].number_input(
            "After-tax deductions", min_value=0.0, value=0.0, step=10.0, help="Examples: advances, benefits recovery, or other non-tax deductions."
        )
        submitted = st.form_submit_button("Calculate payroll", type="primary")

    if submitted:
        payroll_input = {
            "frequency": frequency,
            "hours": hours,
            "rate": rate,
            "employee_id": employee_id,
            "pay_start": pay_start,
            "pay_end": pay_end,
            "pay_date": pay_date,
            "province": province,
            "salary_amount": salary_amount,
            "overtime_hours": overtime_hours,
            "overtime_rate": overtime_rate,
            "stat_pay": stat_pay,
            "sick_pay": sick_pay,
            "vacation_pay": vacation_pay,
            "vac_accrual": vac_accrual,
            "bonus": bonus,
            "reimbursements": reimbursements,
            "federal_claim_amount": federal_claim_amount,
            "provincial_claim_amount": provincial_claim_amount,
            "cpp_exempt": cpp_exempt,
            "ei_exempt": ei_exempt,
            "before_tax_deductions": before_tax_deductions,
            "additional_tax": additional_tax,
            "other_deductions": other_deductions,
        }
        ytd_before = payroll_ytd_before(payroll_register, employee_id)
        payroll_input["ytd_cpp"] = ytd_before["cpp"] + ytd_cpp
        payroll_input["ytd_cpp2"] = ytd_cpp2
        payroll_input["ytd_ei"] = ytd_before["ei"] + ytd_ei
        calc = calculate_payroll(payroll_input)
        register_row = make_payroll_register_row(payroll_input, calc, ytd_before)
        updated_register = pd.concat(
            [payroll_register, pd.DataFrame([register_row], columns=PAYROLL_COLUMNS)],
            ignore_index=True,
        )
        st.session_state["payroll_calc"] = {
            "company": {
                "name": company_name,
                "address": company_address,
                "phone": company_phone,
                "payroll_account": payroll_account,
            },
            "employee": {
                "name": employee_name or "Employee",
                "id": employee_id,
                "position": position,
                "province": province,
                "address": employee_address,
                "hire_date": hire_date,
            },
            "payroll": {
                **payroll_input,
                "pay_start": pay_start,
                "pay_end": pay_end,
                "pay_date": pay_date,
            },
            "calc": calc,
            "updated_register": updated_register,
        }

    saved = st.session_state.get("payroll_calc")
    if saved:
        calc = saved["calc"]
        st.markdown("**Payroll result**")
        metric_cols = st.columns(4)
        metric_cols[0].metric("Gross pay", f"${calc['gross']:,.2f}")
        metric_cols[1].metric("Employee deductions", f"${calc['total_deductions']:,.2f}")
        metric_cols[2].metric("Net pay", f"${calc['net']:,.2f}")
        metric_cols[3].metric("Employer cost add-on", f"${calc['employer_cpp'] + calc['employer_ei']:,.2f}")

        result_df = pd.DataFrame(
            [
                {"Item": "Regular pay", "Amount": calc["regular_pay"]},
                {"Item": "Salary amount", "Amount": saved["payroll"]["salary_amount"]},
                {"Item": "Overtime pay", "Amount": calc["overtime_pay"]},
                {"Item": "Sick pay", "Amount": saved["payroll"]["sick_pay"]},
                {"Item": "Gross pay", "Amount": calc["gross"]},
                {"Item": "Taxable income for tax", "Amount": calc.get("taxable_income_for_tax", calc["gross"])},
                {"Item": "Before-tax deductions", "Amount": -calc.get("before_tax_deductions", 0.0)},
                {"Item": "CPP", "Amount": -calc["cpp"]},
                {"Item": "EI", "Amount": -calc["ei"]},
                {"Item": "Federal tax", "Amount": -calc["tax_fed"]},
                {"Item": "BC tax", "Amount": -calc["tax_prov"]},
                {"Item": "Other deductions", "Amount": -calc["other_deductions"]},
                {"Item": "Reimbursements", "Amount": calc["reimbursements"]},
                {"Item": "Net pay", "Amount": calc["net"]},
                {"Item": "Employer CPP", "Amount": calc["employer_cpp"]},
                {"Item": "Employer EI", "Amount": calc["employer_ei"]},
            ]
        )
        st.dataframe(result_df, use_container_width=True, hide_index=True)

        pdf_payroll = {
            **saved["payroll"],
            "cpp": calc["cpp"],
            "ei": calc["ei"],
            "tax_fed": calc["tax_fed"],
            "tax_prov": calc["tax_prov"],
            "taxable_pay": calc.get("taxable_income_for_tax", calc["gross"]),
            "before_tax_deductions": calc.get("before_tax_deductions", 0.0),
            "total_deductions": calc["total_deductions"],
            "net": calc["net"],
        }
        pdf_bytes = build_payslip_pdf(saved["company"], saved["employee"], pdf_payroll, calc)
        st.download_button(
            "Download payslip PDF",
            data=pdf_bytes,
            file_name=f"{safe_name(saved['employee']['name']) or 'Employee'}_{saved['payroll']['pay_date']}_payslip.pdf",
            mime="application/pdf",
            type="primary",
        )
        excel_payslip_bytes = build_payslip_excel(saved["company"], saved["employee"], pdf_payroll, calc)
        st.download_button(
            "Download payslip Excel",
            data=excel_payslip_bytes,
            file_name=f"{safe_name(saved['employee']['name']) or 'Employee'}_{saved['payroll']['pay_date']}_payslip.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        pd7a_bytes = build_pd7a_excel(saved["company"], saved["payroll"], calc)
        st.download_button(
            "Download PD7A remittance summary",
            data=pd7a_bytes,
            file_name=f"PD7A_Remittance_{saved['payroll']['pay_date']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        register_bytes = export_payroll_register(saved["updated_register"])
        st.download_button(
            "Download updated payroll register",
            data=register_bytes,
            file_name="Payroll_Register_Updated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        with st.expander("Preview payroll register row"):
            st.dataframe(saved["updated_register"].tail(1), use_container_width=True, hide_index=True)
        st.warning("Payroll calculations should be reviewed against CRA PDOC before remitting or filing.")

with mortgage_tab:
    st.subheader("Maximum Mortgage Under GDSR")
    st.caption("Canadian mortgage affordability calculator using GDSR allowance, qualifying rate, property tax, heat, condo fees, amortization, and down payment.")
    components.html(mortgage_calculator_html(), height=920, scrolling=True)

with real_estate_tab:
    st.subheader("Real Estate Investment Agent")
    st.caption("Upload a combined CSV, run the search, then review ranked investment opportunities.")
    st.info(
        "Use one combined CSV with Realtor.ca listing columns. If available, include Zealty columns in the same file "
        "for richer scoring."
    )
    with st.expander("Real estate agent guide"):
        st.markdown(
            """
            **Quick start**

            This tool helps you quickly find better investment opportunities without manually comparing every listing.

            1. Upload one combined CSV under **Zealty-enriched combined CSV**.
            2. Click **Run real estate search**.
            3. Review the Top 10 and Property Database tables.

            **Best one-file workflow**

            Use a **Zealty-enriched combined CSV**. This is one CSV file that contains:

            - Realtor.ca listing columns: address, MLS number, price, property type, beds, baths, square feet, lot size
            - Zealty enrichment columns: sold price history, sale dates, price 1Y/3Y/5Y ago, previous listing prices,
              price change history, days on market, Zealty URL

            Upload this one rich combined file under **Zealty-enriched combined CSV**.

            **Score meaning**

            - Higher score is better.
            - `80-100`: strongest opportunities
            - `60-79`: worth reviewing
            - `40-59`: average / needs more research
            - `0-39`: weaker based on current data

            **Basic Realtor-only score**

            Used when Zealty/rental data is blank. It ranks by lower price per square foot, lower list price,
            larger lot size, more bedrooms, and property type.

            **Full investment score**

            Used when enrichment data is uploaded:

            - Price appreciation: 30%
            - Cash flow: 25%
            - Comparable sales discount: 20%
            - Transit access: 10%
            - School quality: 5%
            - Development potential: 10%

            **Zealty enrichment**

            Put Zealty columns directly inside the combined CSV. No separate optional upload is needed.
            """
        )

    realtor_csv_file = st.file_uploader(
        "Zealty-enriched combined CSV",
        type=["csv"],
        help=(
            "Upload a Realtor.ca listing CSV, or one combined CSV that also includes Zealty columns such as Price 1Y Ago, Price 5Y Ago, "
            "Price Change History, and Days on Market."
        ),
    )

    zealty_url = ""
    zealty_file = None
    rental_file = None
    signal_file = None

    if st.button("Run real estate search", type="primary", use_container_width=True):
        if realtor_csv_file is None:
            st.error("Upload a Zealty-enriched combined CSV first.")
        else:
            try:
                with st.spinner("Searching Realtor.ca and ranking properties..."):
                    property_df, top_df = run_real_estate_search(
                        "",
                        realtor_csv_file,
                        zealty_url.strip(),
                        zealty_file,
                        rental_file,
                        signal_file,
                    )
                st.session_state["real_estate_property_df"] = property_df
                st.session_state["real_estate_top_df"] = top_df
            except Exception as exc:
                st.error(str(exc))

    property_df = st.session_state.get("real_estate_property_df")
    top_df = st.session_state.get("real_estate_top_df")

    if isinstance(property_df, pd.DataFrame) and not property_df.empty:
        metric_cols = st.columns(4)
        metric_cols[0].metric("Listings", f"{len(property_df):,}")
        metric_cols[1].metric("Average list price", f"${property_df['List Price'].mean():,.0f}")
        price_sqft = pd.to_numeric(property_df["Price / Sq Ft"], errors="coerce").dropna()
        metric_cols[2].metric("Median price / sq ft", f"${price_sqft.median():,.0f}" if not price_sqft.empty else "Pending")
        metric_cols[3].metric("Top score", f"{top_df['Investment Score'].max():.0f}" if not top_df.empty else "Pending")

        st.markdown("**Top 10 Investment Opportunities**")
        top_columns = [
            "Address",
            "City",
            "List Price",
            "1-Year Change %",
            "5-Year Change %",
            "Estimated Rent",
            "Estimated Cash Flow",
            "Investment Score",
            "Flags",
            "Notes",
        ]
        visible_top_columns = [column for column in top_columns if column in top_df.columns]
        st.dataframe(top_df[visible_top_columns], use_container_width=True, hide_index=True)

        st.markdown("**Property Database**")
        st.dataframe(property_df, use_container_width=True, hide_index=True)

        download_cols = st.columns(2)
        download_cols[0].download_button(
            "Download property CSV",
            data=property_df.to_csv(index=False).encode("utf-8-sig"),
            file_name="property_database_rows.csv",
            mime="text/csv",
            use_container_width=True,
        )
        download_cols[1].download_button(
            "Download Excel workbook",
            data=real_estate_excel_bytes(
                {
                    "Property Database": property_df,
                    "Top 10 Opportunities": top_df,
                }
            ),
            file_name="property_database.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.info("Run a Realtor.ca search to populate the property database.")

with guide_tab:
    st.subheader("Recommended file names")
    st.caption("You may use any filename. This format makes monthly and annual files easier to sort.")
    st.code(
        "2025-02_BMO_Chequing.pdf\n"
        "2025-02_CIBC_Chequing.pdf\n"
        "2025-02_RBC_Chequing.pdf\n"
        "2025-02_RBC_Visa.pdf\n"
        "2025-02_TD_Chequing.pdf\n"
        "2025-02_Tangerine_Chequing.pdf\n"
        "2025-02_Vancity_Chequing.pdf",
        language="text",
    )
    st.subheader("Review checklist")
    st.markdown(
        "1. Upload statement files and extract Excel.\n"
        "2. Compare total debits, credits, opening balance and closing balance with the statement.\n"
        "3. Keep original PDFs until the annual accounting file is complete.\n"
        "4. Merge the verified monthly workbooks into an annual workbook."
    )
    st.info("Files uploaded to this app are processed for the current session. Configure your hosting provider's privacy and retention settings before using real client statements.")
